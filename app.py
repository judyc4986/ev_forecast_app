import os
import math
from flask import Flask, render_template, request, send_from_directory, flash, redirect, url_for
from openpyxl import load_workbook

# =========================================================
# BASIC FLASK SETUP
# =========================================================
app = Flask(__name__)
app.secret_key = "random_secret_string"

# =========================================================
# PATHS – RENDER & LOCAL FRIENDLY
# =========================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FORMULA_PATH = os.path.join(BASE_DIR, "static", "data", "County-level formula_cleaned.xlsx")
SC_SUMMARY_PATH = os.path.join(BASE_DIR, "static", "data", "supercharger_by_county_summary.xlsx")

CHARTS_DIR = os.path.join(BASE_DIR, "static", "charts")
MAPS_DIR = os.path.join(BASE_DIR, "static", "maps")

# =========================================================
# LOAD FORMULA EXCEL USING OPENPYXL
# =========================================================
if not os.path.exists(FORMULA_PATH):
    raise FileNotFoundError(f"Cannot find Excel file: {FORMULA_PATH}")

wb = load_workbook(FORMULA_PATH)
ws = wb.active

columns = [cell.value for cell in ws[1]]

df_formulas = []
for row in ws.iter_rows(min_row=2, values_only=True):
    row_dict = dict(zip(columns, row))
    row_dict["County_clean"] = str(row_dict["County"]).strip().lower()
    df_formulas.append(row_dict)

# =========================================================
# LOAD SUPERCARGER SUMMARY FILE
# =========================================================
if not os.path.exists(SC_SUMMARY_PATH):
    raise FileNotFoundError(f"Cannot find Supercharger summary: {SC_SUMMARY_PATH}")

wb_sc = load_workbook(SC_SUMMARY_PATH)
ws_sc = wb_sc.active

sc_columns = [cell.value for cell in ws_sc[1]]

supercharger_summary = []
for row in ws_sc.iter_rows(min_row=2, values_only=True):
    row_dict = dict(zip(sc_columns, row))
    row_dict["County_clean"] = str(row_dict["County"]).strip().lower()
    supercharger_summary.append(row_dict)

# =========================================================
# SAFE FORMULA EVALUATOR
# =========================================================
def evaluate_formula(formula_str, x_value):
    if not isinstance(formula_str, str):
        return None

    parts = formula_str.split("=")
    if len(parts) < 2:
        return None

    expr = parts[1].strip()
    expr = expr.replace("^", "**")

    allowed = {
        "x": x_value,
        "exp": math.exp,
        "log": math.log,
        "sqrt": math.sqrt
    }

    try:
        return float(eval(expr, {"__builtins__": {}}, allowed))
    except Exception:
        return None

# =========================================================
# FIND IMAGE FOR COUNTY
# =========================================================
def find_image_for_county(directory, county_name):
    if not os.path.isdir(directory):
        return None

    target = county_name.replace(" ", "_").lower()

    for fname in os.listdir(directory):
        if fname.lower().endswith(".png") and target in fname.lower():
            return fname

    return None

# =========================================================
# ROUTES TO SERVE STATIC CHARTS & MAPS
# =========================================================
@app.route("/chart/<path:filename>")
def serve_chart(filename):
    return send_from_directory(CHARTS_DIR, filename)


@app.route("/map/<path:filename>")
def serve_map(filename):
    return send_from_directory(MAPS_DIR, filename)

# =========================================================
# MAIN APP ROUTE
# =========================================================
@app.route("/", methods=["GET", "POST"])
def index():
    result = None
    chart_filename = None
    map_filename = None
    county_display = None
    supercharger_points = None

    if request.method == "POST":
        county_input = request.form.get("county", "").strip()
        x_input = request.form.get("superchargers", "").strip()

        if not county_input:
            flash("Please enter a county name.")
            return redirect(url_for("index"))

        county_clean = county_input.lower()

        matches = [r for r in df_formulas if r["County_clean"] == county_clean]
        if not matches:
            flash(f"County '{county_input}' not found.")
            return redirect(url_for("index"))

        row = matches[0]
        county_display = row["County"]

        # ------------------------------------------------------
        # Load supercharger count for county
        # ------------------------------------------------------
        sc_match = [r for r in supercharger_summary if r["County_clean"] == county_clean]
        if sc_match:
            supercharger_points = sc_match[0].get("Supercharger_Count", None)

        # ------------------------------------------------------
        # Forecast values if user entered X
        # ------------------------------------------------------
        ev_formula = row.get("EVs_vs_SC_Formula")
        adopt_formula = row.get("Adopt_vs_SC_Formula")

        if x_input:
            try:
                x_val = float(x_input)
            except ValueError:
                flash("Supercharger Points must be a valid number.")
                return redirect(url_for("index"))

            ev_y = evaluate_formula(ev_formula, x_val)
            adopt_y = evaluate_formula(adopt_formula, x_val)

            if ev_y is None or adopt_y is None:
                flash("Could not evaluate formulas for this county.")
                return redirect(url_for("index"))

            result = {
                "x": x_val,
                "ev_y": ev_y,
                "adopt_y": adopt_y,
            }

        # ------------------------------------------------------
        # Find map & chart
        # ------------------------------------------------------
        chart_filename = find_image_for_county(CHARTS_DIR, county_display)
        map_filename = find_image_for_county(MAPS_DIR, county_display)

        if map_filename is None:
            flash(f"No map image found for {county_display}.")
        if chart_filename is None:
            flash(f"No chart image found for {county_display}.")

    return render_template(
        "index.html",
        result=result,
        county=county_display,
        chart_filename=chart_filename,
        map_filename=map_filename,
        supercharger_points=supercharger_points,
    )

# =========================================================
# LOCAL RUN
# =========================================================
if __name__ == "__main__":
    app.run(debug=True)
